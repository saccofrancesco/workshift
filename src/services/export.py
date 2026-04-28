from __future__ import annotations
from datetime import date
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from ..domain.models import Employee, Schedule, Shift
from .view_models import EmployeeWorkloadVM
from .calculations import build_employee_workloads
from .formatting import contrast_text_color, format_month_label


def _solid_fill(color_hex: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=color_hex.replace("#", "").upper())


def _font_color(color_hex: str) -> str:
    return f"FF{color_hex.replace('#', '').upper()}"


def _apply_border(cell) -> None:
    thin: Side = Side(style="thin", color="D7DCE3")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _style_header_row(sheet, row_index: int, columns: int) -> None:
    fill: PatternFill = PatternFill(fill_type="solid", fgColor="EAF0FF")
    for column in range(1, columns + 1):
        cell = sheet.cell(row=row_index, column=column)
        cell.font = Font(bold=True, color="0F172A")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        _apply_border(cell)


def _format_sheet_title(sheet, title: str, subtitle: str, columns: int) -> None:
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=columns)
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=columns)
    title_cell = sheet.cell(row=1, column=1)
    subtitle_cell = sheet.cell(row=2, column=1)
    title_cell.value = title
    subtitle_cell.value = subtitle
    title_cell.font = Font(size=15, bold=True, color="0F172A")
    subtitle_cell.font = Font(size=10, color="475569")
    title_cell.alignment = Alignment(horizontal="left")
    subtitle_cell.alignment = Alignment(horizontal="left")


def _apply_auto_filter(sheet, start_row: int, end_row: int, columns: int) -> None:
    sheet.auto_filter.ref = f"A{start_row}:{get_column_letter(columns)}{end_row}"


def _set_page_setup(sheet) -> None:
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_margins.left = 0.3
    sheet.page_margins.right = 0.3
    sheet.page_margins.top = 0.4
    sheet.page_margins.bottom = 0.4
    sheet.sheet_view.showGridLines = False


def _set_widths(sheet, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def _write_schedule_sheet(sheet, schedule: Schedule, month_date: date) -> None:
    _format_sheet_title(
        sheet,
        f"Workshift Schedule - {format_month_label(month_date)}",
        "Print-friendly planning export",
        6,
    )
    headers_row: int = 4
    headers: list[str] = ["Date", "Weekday", "Employee", "Start", "End", "Hours"]
    for index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=headers_row, column=index, value=header)
        cell.font = Font(bold=True, color="0F172A")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    _style_header_row(sheet, headers_row, len(headers))

    start_row: int = headers_row + 1
    month_key: tuple[int, int] = (month_date.year, month_date.month)
    rows: list[Shift] = sorted(
        [
            shift
            for shift in schedule.shifts
            if (shift.shift_date.year, shift.shift_date.month) == month_key
        ],
        key=lambda shift: (
            shift.shift_date,
            shift.start_time,
            next(
                (
                    employee.full_name.casefold()
                    for employee in schedule.employees
                    if employee.id == shift.employee_id
                ),
                "unknown",
            ),
        ),
    )
    if not rows:
        sheet.cell(row=start_row, column=1, value="No shifts planned for this month.")
        sheet.merge_cells(
            start_row=start_row, start_column=1, end_row=start_row, end_column=6
        )
        sheet.cell(row=start_row, column=1).alignment = Alignment(horizontal="center")
        _set_page_setup(sheet)
        _set_widths(sheet, {"A": 16, "B": 14, "C": 24, "D": 12, "E": 12, "F": 12})
        return

    for index, row in enumerate(rows, start=start_row):
        employee: Employee = next(
            (
                employee
                for employee in schedule.employees
                if employee.id == row.employee_id
            ),
            None,
        )
        date_cell = sheet.cell(row=index, column=1, value=row.shift_date)
        weekday_cell = sheet.cell(
            row=index, column=2, value=row.shift_date.strftime("%A")
        )
        employee_cell = sheet.cell(
            row=index,
            column=3,
            value=employee.full_name if employee else row.employee_name,
        )
        start_cell = sheet.cell(row=index, column=4, value=row.start_time)
        end_cell = sheet.cell(row=index, column=5, value=row.end_time)
        hours_cell = sheet.cell(
            row=index,
            column=6,
            value=(row.end_time.hour + row.end_time.minute / 60)
            - (row.start_time.hour + row.start_time.minute / 60),
        )

        date_cell.number_format = "yyyy-mm-dd"
        start_cell.number_format = "hh:mm"
        end_cell.number_format = "hh:mm"
        hours_cell.number_format = "0.00"

        fill_color: str = employee.color_hex if employee else "#94a3b8"
        employee_cell.fill = _solid_fill(fill_color)
        employee_cell.font = Font(
            color=_font_color(contrast_text_color(fill_color)), bold=True
        )

        for cell in (
            date_cell,
            weekday_cell,
            employee_cell,
            start_cell,
            end_cell,
            hours_cell,
        ):
            cell.alignment = Alignment(vertical="center")
            _apply_border(cell)

    _apply_auto_filter(sheet, headers_row, start_row + len(rows) - 1, len(headers))
    sheet.freeze_panes = "A5"
    _set_page_setup(sheet)
    _set_widths(sheet, {"A": 14, "B": 14, "C": 26, "D": 12, "E": 12, "F": 12})


def _write_workload_sheet(sheet, schedule: Schedule, month_date: date) -> None:
    rows: list[EmployeeWorkloadVM] = build_employee_workloads(schedule, month_date)
    _format_sheet_title(
        sheet,
        f"Workload - {format_month_label(month_date)}",
        "Target hours versus assigned hours",
        5,
    )
    headers_row: int = 4
    headers: list[str] = [
        "Employee",
        "Assigned Hours",
        "Target Hours",
        "Remaining Hours",
        "Utilization",
    ]
    for index, header in enumerate(headers, start=1):
        sheet.cell(row=headers_row, column=index, value=header)
    _style_header_row(sheet, headers_row, len(headers))

    start_row: int = headers_row + 1
    if not rows:
        sheet.cell(row=start_row, column=1, value="No employees available.")
        sheet.merge_cells(
            start_row=start_row, start_column=1, end_row=start_row, end_column=5
        )
        sheet.cell(row=start_row, column=1).alignment = Alignment(horizontal="center")
        _set_page_setup(sheet)
        _set_widths(sheet, {"A": 26, "B": 16, "C": 16, "D": 18, "E": 16})
        return

    for index, row in enumerate(rows, start=start_row):
        employee_cell = sheet.cell(row=index, column=1, value=row.full_name)
        assigned_cell = sheet.cell(row=index, column=2, value=row.assigned_hours)
        target_cell = sheet.cell(row=index, column=3, value=row.target_hours)
        remaining_cell = sheet.cell(row=index, column=4, value=row.remaining_hours)
        utilization_cell = sheet.cell(row=index, column=5, value=row.progress_ratio)

        employee_cell.fill = _solid_fill(row.color_hex)
        employee_cell.font = Font(
            color=_font_color(contrast_text_color(row.color_hex)), bold=True
        )
        for cell in (assigned_cell, target_cell, remaining_cell):
            cell.number_format = "0.00"
        utilization_cell.number_format = "0%"
        for cell in (
            employee_cell,
            assigned_cell,
            target_cell,
            remaining_cell,
            utilization_cell,
        ):
            cell.alignment = Alignment(vertical="center")
            _apply_border(cell)

    _apply_auto_filter(sheet, headers_row, start_row + len(rows) - 1, len(headers))
    sheet.freeze_panes = "A5"
    _set_page_setup(sheet)
    _set_widths(sheet, {"A": 26, "B": 16, "C": 16, "D": 18, "E": 16})
